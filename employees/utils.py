from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from io import BytesIO
from datetime import datetime

def export_employees_pdf(employees):
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    styles = getSampleStyleSheet()
    
    title = Paragraph("Employee Report", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))
    
    data = [['ID', 'Name', 'Department', 'Designation', 'Phone']]
    for emp in employees:
        data.append([
            emp.employee_id,
            emp.user.get_full_name(),
            emp.get_department_display(),
            emp.designation,
            emp.phone
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    doc.build(elements)
    
    buffer.seek(0)
    return buffer

def export_employees_excel(employees):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ['Employee ID', 'Name', 'Email', 'Department', 'Designation', 'Phone', 'Joining Date', 'Salary']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for emp in employees:
        ws.append([
            emp.employee_id,
            emp.user.get_full_name(),
            emp.user.email,
            emp.get_department_display(),
            emp.designation,
            emp.phone,
            emp.date_of_joining.strftime('%Y-%m-%d'),
            float(emp.salary)
        ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def export_attendance_excel(attendance_records):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    
    headers = ['Employee ID', 'Name', 'Date', 'Status', 'Check In', 'Check Out', 'Remarks']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    for record in attendance_records:
        ws.append([
            record.employee.employee_id,
            record.employee.user.get_full_name(),
            record.date.strftime('%Y-%m-%d'),
            record.get_status_display(),
            record.check_in.strftime('%H:%M') if record.check_in else '',
            record.check_out.strftime('%H:%M') if record.check_out else '',
            record.remarks
        ])
    
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def generate_salary_slip_pdf(payroll):
    """Generate professional salary slip PDF"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#2c3e50'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#34495e'),
        spaceAfter=12
    )
    
    # Company Header
    company_name = Paragraph("Your Company Name", title_style)
    elements.append(company_name)
    
    # Salary Slip Title
    slip_title = Paragraph(
        f"Salary Slip - {payroll.month.strftime('%B %Y')}",
        heading_style
    )
    elements.append(slip_title)
    elements.append(Spacer(1, 0.3*inch))
    
    # Employee Information
    emp_data = [
        ['Employee ID:', payroll.employee.employee_id, 'Name:', payroll.employee.user.get_full_name()],
        ['Designation:', payroll.employee.designation, 'Department:', payroll.employee.get_department_display()],
        ['Pay Period:', payroll.month.strftime('%B %Y'), 'Payment Date:', payroll.payment_date.strftime('%d %b %Y') if payroll.payment_date else 'Pending'],
    ]
    
    emp_table = Table(emp_data, colWidths=[1.5*inch, 2*inch, 1.5*inch, 2*inch])
    emp_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (2, 0), (2, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(emp_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Earnings and Deductions Table
    salary_data = [
        ['EARNINGS', 'AMOUNT', 'DEDUCTIONS', 'AMOUNT'],
        ['Basic Salary', f'${payroll.basic_salary:,.2f}', 'Tax', f'${payroll.tax:,.2f}'],
        ['HRA', f'${payroll.hra:,.2f}', 'Provident Fund', f'${payroll.provident_fund:,.2f}'],
        ['Transport Allowance', f'${payroll.transport_allowance:,.2f}', 'Insurance', f'${payroll.insurance:,.2f}'],
        ['Medical Allowance', f'${payroll.medical_allowance:,.2f}', 'Other Deductions', f'${payroll.other_deductions:,.2f}'],
        ['Other Allowances', f'${payroll.other_allowances:,.2f}', '', ''],
        ['Bonus', f'${payroll.bonus:,.2f}', '', ''],
        ['', '', '', ''],
        ['GROSS SALARY', f'${payroll.gross_salary:,.2f}', 'TOTAL DEDUCTIONS', f'${payroll.total_deductions:,.2f}'],
    ]
    
    salary_table = Table(salary_data, colWidths=[2.5*inch, 1.5*inch, 2.5*inch, 1.5*inch])
    salary_table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498db')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data rows
        ('FONTSIZE', (0, 1), (-1, -2), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        
        # Totals row
        ('BACKGROUND', (0, -1), (1, -1), colors.HexColor('#ecf0f1')),
        ('BACKGROUND', (2, -1), (-1, -1), colors.HexColor('#ecf0f1')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 11),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('LINEBELOW', (0, -1), (-1, -1), 2, colors.black),
        
        # Alignment
        ('ALIGN', (1, 1), (1, -1), 'RIGHT'),
        ('ALIGN', (3, 1), (3, -1), 'RIGHT'),
    ]))
    
    elements.append(salary_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Net Salary
    net_salary_data = [
        ['NET SALARY (Gross - Deductions)', f'${payroll.net_salary:,.2f}']
    ]
    
    net_table = Table(net_salary_data, colWidths=[5*inch, 2.5*inch])
    net_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#27ae60')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(net_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # Attendance Summary
    attendance_data = [
        ['Attendance Summary'],
        ['Working Days', str(payroll.working_days)],
        ['Present Days', str(payroll.present_days)],
        ['Absent Days', str(payroll.absent_days)],
        ['Leave Days', str(payroll.leave_days)],
    ]
    
    attendance_table = Table(attendance_data, colWidths=[4*inch, 2*inch])
    attendance_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#95a5a6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('FONTSIZE', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('SPAN', (0, 0), (-1, 0)),
    ]))
    
    elements.append(attendance_table)
    elements.append(Spacer(1, 0.5*inch))
    
    # Footer
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=8,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    
    footer = Paragraph(
        f"This is a computer-generated document. No signature required.<br/>"
        f"Generated on {datetime.now().strftime('%d %B %Y at %H:%M')}",
        footer_style
    )
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer